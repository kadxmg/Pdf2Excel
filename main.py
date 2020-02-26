# -*- coding: utf-8 -*-

import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')  # 编译环境utf8
from glob import glob
import re
import docx
import pdfminer
import math
from docx import Document
from pdfminer.pdfdocument import PDFNoOutlines
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfpage import PDFTextExtractionNotAllowed
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.layout import *
from pdfminer.converter import PDFPageAggregator
##################################这是彩色打印
import ctypes
STD_INPUT_HANDLE = -10
STD_OUTPUT_HANDLE = -11
STD_ERROR_HANDLE = -12


FOREGROUND_BLACK = 0x0
FOREGROUND_BLUE = 0x01  # text color contains blue.
FOREGROUND_GREEN= 0x02  # text color contains green.
FOREGROUND_RED = 0x04  # text color contains red.
FOREGROUND_INTENSITY = 0x08  # text color is intensified.

BACKGROUND_BLUE = 0x10  # background color contains blue.
BACKGROUND_GREEN= 0x20  # background color contains green.
BACKGROUND_RED = 0x40  # background color contains red.
BACKGROUND_INTENSITY = 0x80 # background color is intensified.
# 上面这一大段都是在设置前景色和背景色，其实可以用数字直接设置，我的代码直接用数字设置颜色


class Color:
    std_out_handle = ctypes.windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)

    def set_cmd_color(self, color, handle=std_out_handle):
        bool = ctypes.windll.kernel32.SetConsoleTextAttribute(handle, color)
        return bool

    def reset_color(self):
        self.set_cmd_color(FOREGROUND_RED | FOREGROUND_GREEN | FOREGROUND_BLUE | FOREGROUND_INTENSITY)
        # 初始化颜色为黑色背景，纯白色字，CMD默认是灰色字体的

    def print_red_text(self, print_text):
        self.set_cmd_color(4 | 8)
        print(print_text)
        self.reset_color()
        # 红色字体

    def print_green_text(self, print_text):
        self.set_cmd_color(FOREGROUND_GREEN | FOREGROUND_INTENSITY)
        # c = raw_input(print_text.encode('gbk'))
        # c = raw_input(print_text)
        print(print_text)
        self.reset_color()
        # return c

    def print_yellow_text(self, print_text):
        self.set_cmd_color(6 | 8)
        print(print_text)
        self.reset_color()
        # 黄色字体

    def print_blue_text(self, print_text):
        self.set_cmd_color(1 | 10)
        print(print_text)
        self.reset_color()
        # 蓝色字体


clr = Color()
clr.set_cmd_color(FOREGROUND_RED | FOREGROUND_GREEN | FOREGROUND_BLUE | FOREGROUND_INTENSITY)
# clr.print_red_text('red')
# clr.print_green_text("green")
# clr.print_blue_text('blue')
# clr.print_yellow_text('yellow')
##########################################


PROJECT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))

import fnmatch

def get_filename_by_path(path, request_name, forbid_word=''):
    matches = []
    for root, dirnames, filenames in os.walk(path):
        for filename in fnmatch.filter(filenames, request_name):
            if(len(forbid_word) == 0 or filename.count(forbid_word) == 0):
                matches.append(os.path.join(root, filename))
    print("get_filename_by_path %s %s %s" % (path,request_name,forbid_word))
    print("get_filename_by_path %s" % matches)
    return matches

    #searched_filename = glob(path)
    #return [i for i in searched_filename if forbid_word not in i ]


def read_pdf(filename):
    results = []

    fp = open(filename, 'rb')
    parser = PDFParser(fp)
    document = PDFDocument(parser)
    if not document.is_extractable:
        raise PDFTextExtractionNotAllowed
    print(u'定位并解析数据')
    content = ""

    try:
        outlines = document.get_outlines()
        for (level, title, dest, a, se) in outlines:
            print (level, title)
    except PDFNoOutlines:
        print(u'没有大纲')

    #  创建一个PDF资源管理器对象来存储共赏资源
    rsrcmgr=PDFResourceManager()
    # 设定参数进行分析
    laparams=LAParams()
    # 创建一个PDF设备对象
    # device=PDFDevice(rsrcmgr)
    device=PDFPageAggregator(rsrcmgr,laparams=laparams)
    # 创建一个PDF解释器对象
    interpreter=PDFPageInterpreter(rsrcmgr,device)

    current_section = "" # use for save tag
    last_section = "" # use for save title
    current_char_size = 0;
    current_fontname = "";
    debug_line = False
    debug_section = False
    debug_title = True
    # 处理每一页
    for page in PDFPage.create_pages(document):
        interpreter.process_page(page)
        # 接受该页面的LTPage对象
        layout = device.get_result()
        for listItem in layout:
            if(isinstance(listItem,LTTextBox)):
                # print("LTTextBox: %s" % listItem.get_text().encode('utf-8')+'\n')
                for textLine in listItem:
                    if (isinstance(textLine, LTTextLine)):
                        if debug_line:
                            if (len(textLine.get_text().strip())):
                                print("LTTextLine: %s" % textLine.get_text().encode('utf-8') + '\n')
                        for char in textLine:
                            if (isinstance(char, LTChar)):
                                #if(len(char.get_text().strip()) > 0):
                                #    print(" LTChar: %s size:%s font: %s " %  (char,char.size,char.fontname))

                                # new section ?
                                if(abs(char.size - current_char_size) > 0.00001 and len(current_section.strip()) > 0 or current_fontname != char.fontname):
                                    #print("    size chang from : %s  to: %s %s" % (char.size, current_char_size,char))
                                    try:
                                        if debug_section:
                                            print(u"current_section : %s" % (current_section))
                                            print(u"last_section : %s" % (last_section))
                                    except UnicodeEncodeError:
                                        clr.print_red_text("UnicodeEncodeError")

                                    # print("section: %s : %f - %f" %  (current_section,current_char_size,char.size))
                                    # print(" %s - %s " %  (type(current_char_size),type(char.size)))
                                    current_section = current_section.strip()

                                    if current_section.startswith("<KPOC-REQ") or  current_section.startswith("KPOC-REQ"):
                                        # this is req tag
                                        if current_section.startswith("KPOC-REQ") :
                                            current_section = "<" + current_section
                                        if current_section.count(".......") > 0 :
                                            # index , new result
                                            result = {
                                                "Section" : "",
                                                'Title': None,
                                                'ReqId': None,
                                                'Content': "",
                                            }

                                            ReqId_end = current_section.find('>')
                                            Title_end = current_section.find('............')
                                            ReqId = current_section[0:ReqId_end + 1]
                                            Title = current_section[ReqId_end + 1:Title_end]
                                            Section = last_section
                                            result["Title"] = Title.strip()
                                            result["ReqId"] = ReqId.strip()
                                            result["Section"] = Section.strip()
                                            results.append(result)  # save the result
                                            if debug_title:
                                                print(u"Text: %s" % (current_section))
                                                print(u"Section: %s" % (Section))
                                                print(u"ReqId: %s" % (ReqId))
                                                print(u"Title: %s" % (Title))
                                                print(u"\n")
                                        else:
                                            # not index , old result
                                            ReqId_end = current_section.find('>')
                                            ReqId = current_section[0:ReqId_end + 1]
                                            for result in results:
                                                if result["ReqId"] == ReqId:
                                                    Content_Begin = current_section.find('> ') + 2
                                                    result["Content"] = current_section[Content_Begin:]
                                                    #exit(0)

                                        result = None; # clean result
                                        current_section = "" # remove the title from next content
                                        if debug_section:
                                            print(u"current_section clear for ReqId")
                                    else:
                                        # endif with current_section.startswith
                                        # after all , when a new section is found , we should clean this
                                        if(len(current_section.strip())):
                                            last_section = current_section # section number maybe here
                                        current_section = ""
                                        if debug_section:
                                            print(u"current_section clear for Not Found ReqId")
                                            #print(u"last_section : %s" % (last_section))

                                # same section
                                current_section = current_section + char.get_text()
                                current_char_size = char.size
                                current_fontname = char.fontname

    #save the last para
    if len(results) > 0: # skip the first one
        content = content.strip()
        results[-1]["Content"] = content # save the last one content
    content = "" # empty the content

    fp.close()

    print("count %d" % len(results))
    return results

def write_excel(excel_name, result_dicts):
    from openpyxl.workbook import Workbook
    
    from openpyxl.styles import Alignment
    alignment = Alignment(
        wrap_text = True, # 自动换行
    )

    #ExcelWriter,里面封装好了对Excel的写操作
    from openpyxl.writer.excel import ExcelWriter

    #get_column_letter函数将数字转换为相应的字母，如1-->A,2-->B
    from openpyxl.utils  import get_column_letter

    from openpyxl.reader.excel import load_workbook

    if os.path.isfile(excel_name):
        # #读取excel2007文件
        wb = load_workbook(excel_name)
    else:
        #新建一个workbook
        wb = Workbook()

    #设置文件输出路径与名称
    dest_filename = excel_name

    
    # # 获取第一个sheet

    ws = wb.get_active_sheet()
    if ws != None:
        wb.remove_sheet(ws)
       
    ws = wb.create_sheet('Sheet1')


    #第一个sheet是ws
    # ws = wb.worksheets[0]

    # #设置ws的名称
    # ws.title = "sheet1"

    line = 1
    print(u'定位写入坐标')
    while ws.cell(line,1).value:
        # print(ws.cell("A%s" % line).value)
        line += 1
    print(u'从第%s行开始写入' % line)


    #Title
    col = 1
    ws.cell(line, col).value = u'Chapter'
    col = col + 1
    ws.cell(line, col).value = u'ReqId'
    col = col + 1
    ws.cell(line, col).value = u'Title'
    col = col + 1
    ws.cell(line, col).value = u'Content'

    ws.column_dimensions['D'].width = 50.0
    line += 1
    
    for i, result in enumerate(result_dicts):
        #print(u'正在写入第%s条数据到excel' % (i+1))
        try:
            print(u'正在写入 %s' % result['ReqId'])
        except  UnicodeEncodeError:
            clr.print_red_textprint("UnicodeEncodeError")
        ws.cell(line, 2).value = result['ReqId']
        ws.cell(line, 3).value = result['Title']
        if(len(result['Content']) > 300):
            ws.cell(line, 4).value = result['Content'][:300] + " ... "
        else:
            ws.cell(line, 4).value = result['Content']
        ws.cell(line, 4).alignment = alignment
        line += 1

    #最后保存文件
    wb.save(filename=dest_filename)
    
def main():
    print(u'开始执行')
    print(u'从input文件夹查找文件')
    result_dicts = []

    filenames = get_filename_by_path('input','*.pdf')
    for filename in filenames:
        print(u'读取文件：')
        clr.print_blue_text(os.path.basename(filename))

        results = read_pdf(filename)
        # print(len(results))
        # add for each one
        for result in results:
            result_dicts.append(result)

    # save the result
    save_filename = 'output/output.xlsx'
    # save_filename = 'output/output%s.xlsx' % int(time.time())
    write_excel(save_filename, result_dicts)
    print(u'执行完毕，文件保存至')
    clr.print_blue_text(save_filename)
    # print(save_filename)
    print(u'敲击回车结束运行')
    raw_input()

if __name__ == '__main__':
    main()
